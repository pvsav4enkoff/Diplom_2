from datetime import date, datetime

from django.core.paginator import Paginator
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.db.models import Q
from plant.forms import *
from .models import Equipment
from openpyxl import Workbook
from django.contrib.auth import get_user_model
from django.contrib.auth.hashers import check_password
from django.contrib.auth.decorators import login_required

from django.contrib.auth import authenticate, login, logout
from .forms import LoginForm
from dateutil import parser




def login_view(request):
    error = ''
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('plant')
            else:
                error = 'Не верный логин или пароль.'
    else:
        form = LoginForm()
    return render(request, 'login.html', {'form': form, 'error': error})


def verify_admin_password(password):
    # Получаем модель пользователя
    User = get_user_model()
    # Получаем администратора
    admin_users = User.objects.filter(is_superuser=True)
    admin_user = admin_users.first()
    return check_password(password, admin_user.password)


def complete_repair(request, pk):
    info = {}
    error = None
    repair = Repair.objects.get(id=pk)
    if request.method == 'POST':
        form = CompleteRepairForm(request.POST, instance=repair)
        password = request.POST['password']  # получаем пароль из POST-запроса
        if form.is_valid():
            if not verify_admin_password(password):
                # print('Пароль администратора не правильный')
                error = 'Пароль администратора не правильный'
                info['error'] = error
            else:
                if repair.end_date.date() < repair.start_date.date():
                    error = f'Дата окончания не может быть меньше даты начала'
                    info['error'] = error
                    # print('Дата окончания не может быть меньше даты начала')
                elif repair.end_date.date() > date.today():
                    error = 'Дата окончания не может быть больше текущей.'
                    info['error'] = error
                    # print('Дата окончания не может быть больше текущей.')
                else:
                    repair.status = 'Завершено'
                    form.save()
                    # print('Задача успешно завершена')
                    return redirect('repairs_list')
        else:
            print(form.errors)
    else:
        form = CompleteRepairForm(instance=repair)
    info = {'repair': repair, 'form': form, 'error': error}
    return render(request, 'complete_repair.html', info)


def complete_task(request, pk):
    info = {}
    error = None
    task = Task.objects.get(id=pk)
    if request.method == 'POST':
        form = CompleteTaskForm(request.POST, instance=task)
        password = request.POST['password']  # получаем пароль из POST-запроса
        description = request.POST['description']  # получаем пароль из POST-запроса
        if form.is_valid():
            if not verify_admin_password(password):
                # print('Пароль администратора не правильный')
                error = 'Пароль администратора не правильный'
                info['error'] = error
            else:
                if task.end_date.date() < task.start_date.date():
                    error = f'Дата окончания не может быть меньше даты начала'
                    info['error'] = error
                    # print('Дата окончания не может быть меньше даты начала')
                elif task.end_date.date() > date.today():
                    error = 'Дата окончания не может быть больше текущей.'
                    info['error'] = error
                    # print('Дата окончания не может быть больше текущей.')
                elif len(description) == 0:
                    error = 'Описание работ не долно быть пустым.'
                    info['error'] = error
                else:
                    task.status = 'Завершено'
                    task.description = description
                    form.save()
                    # print('Задача успешно завершена')
                    return redirect('task_list')
        else:
            print(form.errors)
    else:
        form = CompleteTaskForm(instance=task)

    info = {'task': task, 'form': form, 'error': error}
    return render(request, 'complete_task.html', info)

def export_to_excel_task(request):

                if request.method != 'GET':
                    return HttpResponse('Неправильный метод запроса')

                start_date = request.GET.get('start_datef')
                start_date2 = request.GET.get('start_date2f')

                if start_date is None or start_date2 is None:
                    return HttpResponse('Необходимые параметры не найдены')

                # Создаем новый workbook
                wb = Workbook()
                ws = wb.active

                # Добавляем заголовки столбцов
                ws['A1'] = 'Оборудование'
                ws['B1'] = 'Объект'
                ws['C1'] = 'Исполнитель'
                ws['D1'] = 'Руководитель работ'
                ws['E1'] = 'Описание'
                ws['F1'] = 'Дата начала'
                ws['G1'] = 'Дата окончания'
                ws['H1'] = 'Статус'
                print(start_date,start_date2)
                # Добавляем данные в таблицу
                posts = Task.objects.all()
                if start_date :
                    posts = posts.filter(start_date__gte=start_date, start_date__lte=start_date2)

                for i, post in enumerate(posts, start=2):
                    ws[f'A{i}'] = post.repair.equipment.full_info
                    ws[f'B{i}'] = post.task_object.name
                    ws[f'C{i}'] = post.employee.name
                    ws[f'D{i}'] = post.employee_status
                    ws[f'E{i}'] = post.description
                    ws[f'F{i}'] = post.start_date.strftime('%d.%m.%Y %H:%M')
                    ws[f'G{i}'] = post.end_date.strftime('%d.%m.%Y %H:%M') if post.end_date else ''
                    ws[f'H{i}'] = post.status

                # Сохраняем workbook в файл
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': 'attachment; filename="posts.xlsx"'}
                )
                wb.save(response)

                return response


def view_plant(request):
    header = "Учет регламентных работ и ремонта. "
    context = {'header': header}
    return render(request, 'plant.html', context)


def view_plant_employee(request):
    header = "Работники"
    # employee = Employee.objects.all().order_by('id')
    employee = Employee.objects.select_related('brigade').all().order_by('id')
    # # создаем пагинатор
    posts_per_page = request.GET.get('posts_per_page', 10)
    paginator = Paginator(employee, posts_per_page)  # 10 постов на странице

    # получаем номер страницы, на которую переходит пользователь
    page_number = request.GET.get('page', 1)

    # получаем посты для текущей страницы
    page_posts = paginator.get_page(page_number)
    column_headers = ['ФИО', 'Бригада', 'Должность', 'Руководитель работ', 'Активность']
    context = {'header': header, 'page_post': page_posts, 'column_headers': column_headers}
    return render(request, 'employee.html', context)


def view_plant_repair(request):
    header = "Участки"
    sort_by = request.GET.get('sort_by', 'id')  # параметр сортировки
    order_by = request.GET.get('order_by', 'desc')  # порядок сортировки

    start_date = request.GET.get('start_date')
    start_date2 = request.GET.get('start_date2')

    if order_by == 'desc':
        sort_by = '-' + sort_by

    repair = Repair.objects.select_related('brigade', 'equipment').all().order_by(sort_by)
    if start_date and start_date2:
        repair = repair.filter(start_date__gte=start_date, start_date__lte=start_date2)

    tasks = Task.objects.select_related('task_object', 'repair', 'employee').all()

    # # создаем пагинатор
    posts_per_page = request.GET.get('posts_per_page', 10)
    paginator = Paginator(repair, posts_per_page)  # 10 постов на странице

    # получаем номер страницы, на которую переходит пользователь
    page_number = request.GET.get('page', 1)

    # получаем посты для текущей страницы
    page_posts = paginator.get_page(page_number)
    column_headers = ['Площадка', 'Участок', 'Бригада', 'Начало ремонта', 'Окончание ремонта', 'Статус']
    column_headers_task = ['Исполнитель', 'Предмет ремонта', 'Начало', 'Завершение']
    # context = {'header': header, 'page_post': page_posts}

    context = {'header': header, 'page_post': page_posts}
    context['column_headers'] = column_headers
    context['column_headers_task'] = column_headers_task
    context['tasks'] = tasks
    return render(request, 'repair.html', context)


def sort_view(request):
    # Допустимые значения для sort_by
    allowed_sort_by = ['start_date', 'end_date']  # замените на актуальные названия полей модели Task

    sort_by = request.GET.get('sort_by', 'start_date')
    order_by = request.GET.get('order_by', 'asc')

    # Проверка на допустимые значения
    if sort_by not in allowed_sort_by:
        sort_by = 'start_date'
    if order_by not in ['asc', 'desc']:
        order_by = 'asc'

    order = '' if order_by.lower() == 'asc' else '-'
    # print(f'{order}{sort_by}')
    return Task.objects.order_by(f'{order}{sort_by}')


def view_plant_task(request):
    header = "Задачи ремонта"

    sort_by = request.GET.get('sort_by', 'id')  # параметр сортировки
    order_by = request.GET.get('order_by', 'desc')  # порядок сортировки

    start_date = request.GET.get('start_date')
    start_date2 = request.GET.get('start_date2')
    if order_by == 'desc':
        sort_by = '-' + sort_by

    task = Task.objects.select_related('task_object', 'repair', 'employee').all().order_by(sort_by)

    if start_date and start_date2:
        task = task.filter(start_date__gte=start_date, start_date__lte=start_date2)

    # сортируем задачи
    task = task.order_by(sort_by)

    # # создаем пагинатор
    try:
        posts_per_page = int(request.GET.get('posts_per_page', 10))
    except ValueError:
        posts_per_page = 10

    paginator = Paginator(task, posts_per_page)  # 10 постов на странице

    # получаем номер страницы, на которую переходит пользователь
    page_number = request.GET.get('page', 1)

    # получаем посты для текущей страницы
    page_posts = paginator.get_page(page_number)
    column_headers = ['Участок', 'Объект', 'Сотрудник', 'Руководитель работ', 'Дата начала', 'Дата окончания', 'Статус']

    context = {'header': header, 'page_post': page_posts}
    context['column_headers'] = column_headers
    return render(request, 'task.html', context)


def view_plant_equipment(request):
    header = "Площадки завода"
    equipment = Equipment.objects.select_related('segment').prefetch_related('location').all().order_by('id')
    # # создаем пагинатор
    posts_per_page = request.GET.get('posts_per_page', 10)
    paginator = Paginator(equipment, posts_per_page)  # 10 постов на странице

    # получаем номер страницы, на которую переходит пользователь
    page_number = request.GET.get('page', 1)

    # получаем посты для текущей страницы
    page_posts = paginator.get_page(page_number)
    column_headers = ['Площадка', 'Участок', 'Комментарии']
    context = {'header': header, 'page_post': page_posts}
    context['column_headers'] = column_headers
    return render(request, 'equipment.html', context)


def sign_up_by_plant(request):
    error = None
    brigades = []
    positions = []
    br = Brigade.objects.all().values_list('name', flat=True)
    for b in br:
        brigades.append(b)
    pos = Position.objects.all().values_list('name', flat=True)
    for p in pos:
        positions.append(p)
    info = {'good': None, 'error': None, 'brigades': brigades, 'positions': positions}
    if request.method == "POST":
        form = EmployeeRegistr(request.POST)
        if form.is_valid():
            username = form.cleaned_data['name']
            users = Employee.objects.all().values_list('name', flat=True)
            if username in users:
                error = 'Такой сотрудник уже существует.'
            else:
                position = form.cleaned_data['position']
                boss = form.cleaned_data['boss']
                brigade = form.cleaned_data['brigade']
                br = Brigade.objects.get(name=brigade).id
                # print('выбор',str(br),error,username,users)
                Employee.objects.create(name=username, position=position, boss=boss, activ=True, brigade_id=br)
                good = f'Приветствуем, {username}!'
                info['good'] = good
                info['error'] = error
        else:
            error = 'Форма не валидна.'
        info['error'] = error
    else:
        form = EmployeeRegistr()
    info['form'] = form
    return render(request, 'registration_page.html', info)


def registration_equipment(request):
    error = None
    locations_id = []
    segments_id = []
    locations_n = []
    segments_n = []
    locations = []


    locat = LocationZone.objects.all().values_list('zone_location', flat=True)
    for b in locat:
        locations_n.append(b)
    locat = LocationZone.objects.all().values_list('id', flat=True)
    for b in locat:
        locations_id.append(b)
    locations = zip(locations_id, locations_n)
    # print(locations)
    seg = Segment.objects.all().values_list('name', flat=True)
    for p in seg:
        segments_n.append(p)
    seg = Segment.objects.all().values_list('id', flat=True)
    for p in seg:
        segments_id.append(p)
    segments = zip(segments_id, segments_n)
    info = {'good': None, 'error': None, 'segments': segments, 'locations': locations}
    if request.method == "POST":
        form = EquipmentForm(request.POST)
        if form.is_valid():
            location = form.cleaned_data['location']
            segment = form.cleaned_data['segment']
            li = LocationZone.objects.get(zone_location=location).id
            si = Segment.objects.get(name=segment).id
            loc = Equipment.objects.all().values_list('location', flat=True)
            seg = Equipment.objects.all().values_list('segment', flat=True)
            loc_seg = zip(loc, seg)

            for l, s in loc_seg:
                if (str(li).lower() + str(si).lower()) == (str(l).lower() + str(s).lower()):
                    error = 'Такой объект уже существует.'
                    break
            if error is None:
                description = form.cleaned_data['description']
                Equipment.objects.create(description=description, location_id=li, segment_id=si)
                good = f'Объект {location}.{segment} успешно внесен!'
                info['good'] = good
                info['error'] = error
        else:
            error = 'Форма не валидна.'
        info['error'] = form.errors
    else:
        form = EmployeeRegistr()
    info['form'] = form
    return render(request, 'registration_equipment.html', info)


def registration_repair(request):
    error = None
    brigades_id = []
    brigades_n = []

    status = ['В работе', 'Завершен']

    eq_loc = Equipment.objects.all().order_by('id').values_list('location', flat=True)
    br_full = Brigade.objects.all().order_by('id')

    seg = Brigade.objects.all().values_list('name', flat=True)
    for p in seg:
        brigades_n.append(p)
    seg = Brigade.objects.all().values_list('id', flat=True)
    for p in seg:
        brigades_id.append(p)
    brigades = zip(brigades_id, br_full)
    info = {'good': None, 'error': None, 'brigades': brigades, 'status': status}
    if request.method == "POST":

        form = RepairForm(request.POST)
        if form.is_valid():
            equipment = form.cleaned_data['equipment']
            brigade = form.cleaned_data['brigade']
            start_date = form.cleaned_data['start_date']
            # end_date = form.cleaned_data['end_date']
            status = status[0]
            e_id = Repair.objects.all().values_list('equipment_id', 'brigade_id', 'end_date', 'start_date', flat=False)
            z = list(e_id)
            for r in z:
                if (equipment.id, brigade.id, None) == (r[0], r[1], r[2]):
                    # print(equipment.id, equipment.full_info, r[0], brigade.id, brigade.name, r[1], r)
                    error = f'Ремонт уже существует от {r[3]}.'
            if error is None:
                Repair.objects.create(brigade_id=brigade.id, equipment_id=equipment.id, start_date=start_date,
                                      status=status)
                good = f'Ремонт {equipment.full_info}.{brigade} успешно внесен!'
                info['good'] = good
                info['error'] = error
        else:
            error = 'Форма не валидна.'
    else:
        form = RepairForm()

    eq_i = Equipment.objects.all().order_by('id').values_list('id', flat=True)
    eq_full = Equipment.objects.all().order_by('id')

    equip_i = []
    for b in eq_i:
        equip_i.append(b)
    equip_l = []
    for b in eq_loc:
        equip_l.append(b)
    eq = zip(equip_i, eq_full)
    info['eq'] = eq
    # # Вывести результаты
    for equipment in eq_full:
        print(equipment.full_info)

    info['error'] = error
    info['form'] = form
    return render(request, 'registration_repair.html', info)


def registration_task(request):
    error = None
    status = ['В работе', 'Завершен']

    task_objects_n = []
    task_objects_id = []
    task_objects_full = Task_Object.objects.all().order_by('id')
    seg = Task_Object.objects.all().order_by('id').values_list('id', flat=True)
    for p in seg:
        task_objects_id.append(p)

    employee_id = []
    employee_n = []
    empl_full = Employee.objects.filter(activ=True).order_by('id')

    seg = Employee.objects.filter(activ=True).order_by('id').values_list('id', flat=True)
    for p in seg:
        employee_id.append(p)

    repair_id = Repair.objects.filter(Q(end_date=None)).order_by('id').values_list('id', flat=True)
    rep_full = Repair.objects.filter(Q(end_date=None)).order_by('id')
    rep_id = []
    for p in repair_id:
        rep_id.append(p)

    task_objects = zip(task_objects_id, task_objects_full)
    repairs = zip(rep_id, rep_full)
    employees = zip(employee_id, empl_full)

    info = {'good': None, 'error': None, 'repairs': repairs, 'employees': employees, 'task_objects': task_objects,
            'status': status}
    if request.method == "POST":

        form = TaskForm(request.POST)
        if form.is_valid():
            repair = form.cleaned_data['repair']
            task_object = form.cleaned_data['task_object']
            employee = form.cleaned_data['employee']
            employee_status = form.cleaned_data['employee_status']
            start_date = form.cleaned_data['start_date']
            status = status[0]
            t_id = Task.objects.all().values_list('employee_id', 'task_object_id', 'repair_id', 'end_date',
                                                  'start_date', flat=False)
            z = list(t_id)
            for r in z:
                if (employee.id, task_object.id, repair.id, None) == (r[0], r[1], r[2], r[3]):
                    error = f'Задача уже существует от {r[4]}.'
                    info['error'] = error

            if start_date.date() > date.today():
                error = f'Дата начала не может быть больше текущей.'
                info['error'] = error

            if error is None:
                Task.objects.create(employee_id=employee.id, employee_status=employee_status,
                                    task_object_id=task_object.id, repair_id=repair.id,
                                    start_date=start_date, status=status)
                good = f'Ремонт {repair.select_info}.{employee} успешно внесен!'
                info['good'] = good
                info['error'] = error
        else:
            error = 'Форма не валидна.'
            if error is None:
                info['error'] = form.errors
    else:
        form = RepairForm()

    info['current_date'] = date.today()
    info['form'] = form
    return render(request, 'registration_task.html', info)


def delete_post(request, pk):
    if request.method == 'POST':
        post = get_object_or_404(Employee, pk=pk)
        post.delete()
        return redirect('employee.html')


def edit_employee(request, pk):
    error = None
    brigades = []
    brigades_id = []
    positions = []
    br = Brigade.objects.all().values_list('name', flat=True).order_by('id')
    for b in br:
        brigades.append(b)
    br = Brigade.objects.all().values_list('id', flat=True).order_by('id')
    for b in br:
        brigades_id.append(b)
    brigades_all = zip(brigades_id, brigades)
    pos = Position.objects.all().values_list('name', flat=True)
    for p in pos:
        positions.append(p)
    info = {'error': error, 'brigades': brigades, 'brigades_id': brigades_id, 'positions': positions,
            'brigades_all': brigades_all}
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        form = EmployeeForm(request.POST, instance=employee)
        if form.is_valid():
            info['good'] = f'{employee.name} изменен.'
            employee.save()
            return redirect('employee_main')
        else:
            print('Форма не валидна.')
    else:
        form = EmployeeForm(instance=employee)
    info['form'] = form
    info['name'] = employee.name
    info['position'] = employee.position
    info['boss'] = employee.boss
    info['activ'] = employee.activ
    info['selected_brigade'] = employee.brigade_id

    return render(request, 'edit_employee.html', info)


